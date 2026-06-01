#!/usr/bin/env python3
"""
update_data.py — Sincroniza los datos de los archivos Excel con index.html.

Ejecutar cada vez que se actualice alguno de los Excel fuente:
    raw/padb_eje_cafetero.xlsx   → acciones del plan  (DEFAULT_DATA)
    raw/actores.xlsx             → organizaciones (ORGS_DATA) y personas (PERSONAS_DATA)

Uso:
    python update_data.py

Requisitos:
    pip install openpyxl
"""

import csv
import openpyxl
import json
import re
import sys
import io
import os

# Fix encoding para terminales Windows
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'buffer'):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
HTML_FILE    = os.path.join(SCRIPT_DIR, 'index.html')
PADB_FILE    = os.path.join(SCRIPT_DIR, 'raw', 'padb_eje_cafetero.xlsx')
ACTORES_FILE = os.path.join(SCRIPT_DIR, 'raw', 'actores.xlsx')
RED_CSV      = os.path.join(SCRIPT_DIR, 'raw', 'resultados_red.csv')


# ─── helpers ────────────────────────────────────────────────────────────────

def cs(val):
    """Cell value → string (preserva saltos de línea de celdas multi-línea)."""
    return '' if val is None else str(val)

def j(v):
    """json.dumps sin escape de caracteres no-ASCII."""
    return json.dumps(v, ensure_ascii=False)

def parse_coord(val):
    """'lat, lng' string → [float, float] o None."""
    if not val:
        return None
    try:
        parts = str(val).split(',')
        if len(parts) == 2:
            return [float(parts[0].strip()), float(parts[1].strip())]
    except (ValueError, IndexError):
        pass
    return None

def parse_keywords(val):
    """Keywords separados por '-' o ',' → lista de strings."""
    if not val:
        return []
    return [k.strip() for k in re.split(r'[-,]', str(val)) if k.strip()]


# ─── lectura de Excel ────────────────────────────────────────────────────────

def read_padb():
    wb = openpyxl.load_workbook(PADB_FILE)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        has_content = False
        for c, h in enumerate(headers, 1):
            if h is None:
                continue
            val = cs(ws.cell(r, c).value)
            row[h] = val
            if val:
                has_content = True
        if has_content:
            rows.append(row)
    return rows


def read_actores():
    wb = openpyxl.load_workbook(ACTORES_FILE)

    ws_o = wb['organizaciones']
    oh = [ws_o.cell(1, c).value for c in range(1, ws_o.max_column + 1)]
    orgs = []
    for r in range(2, ws_o.max_row + 1):
        row = {h: ws_o.cell(r, c).value for c, h in enumerate(oh, 1)}
        if row.get('organización'):
            orgs.append(row)

    ws_p = wb['personas']
    ph = [ws_p.cell(1, c).value for c in range(1, ws_p.max_column + 1)]
    personas = []
    for r in range(2, ws_p.max_row + 1):
        row = {h: ws_p.cell(r, c).value for c, h in enumerate(ph, 1)}
        if row.get('Nombre completo del participante'):
            personas.append(row)

    return orgs, personas


# ─── conversión a JavaScript ─────────────────────────────────────────────────

def padb_to_js(rows):
    lines = [f"  {j(row)}" for row in rows]
    return "[\n" + ",\n".join(lines) + "\n]"


def orgs_to_js(orgs):
    lines = []
    for i, o in enumerate(orgs):
        coord    = parse_coord(o.get('coordenada'))
        coord_js = f"[{coord[0]},{coord[1]}]" if coord else "null"
        kws      = parse_keywords(o.get('keywords', ''))
        link     = o.get('link')
        biowiki  = o.get('biowiki')

        props = [
            f"id:{i}",
            f"org:{j(cs(o.get('organización','')))}",
            f"bio:{j(cs(o.get('bioeconomia','')))}",
            f"keywords:{j(kws)}",
            f"link:{j(cs(link) if link else None)}",
            f"biowiki:{j(cs(biowiki) if biowiki else None)}",
            f"coord:{coord_js}",
            f"foto:{j(cs(o.get('foto','')))}",
        ]
        lines.append("  {" + ",".join(props) + "}")
    return "[\n" + ",\n".join(lines) + "\n]"


def personas_to_js(personas):
    lines = []
    for p in personas:
        red = p.get('Red Social')
        props = [
            f"nombre:{j(cs(p.get('Nombre completo del participante','')))}",
            f"perfil:{j(cs(p.get('Perfil del participante','')))}",
            f"red:{j(cs(red) if red else None)}",
            f"email:{j(cs(p.get('Correo electrónico','')))}",
            f"departamento:{j(cs(p.get('Departamento','')))}",
            f"cargo:{j(cs(p.get('Cargo','')))}",
            f"genero:{j(cs(p.get('Género','')))}",
            f"org:{j(cs(p.get('Organizacion','')))}",
        ]
        lines.append("  {" + ",".join(props) + "}")
    return "[\n" + ",\n".join(lines) + "\n]"


# ─── lectura de CSV red ──────────────────────────────────────────────────────

def read_red_csv():
    respuestas = []
    with open(RED_CSV, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            colabs = [c.strip() for c in row.get('Entidades_Colaboradoras','').split(',') if c.strip()]
            deptos = [d.strip() for d in row.get('Departamentos','').split(',')           if d.strip()]
            try:
                municipios = json.loads(row.get('Municipios_Por_Depto','{}') or '{}')
            except Exception:
                municipios = {}
            try:
                sectores = json.loads(row.get('Sectores_Por_Municipio','{}') or '{}')
            except Exception:
                sectores = {}
            respuestas.append({
                'nombre':        row.get('Nombre','').strip(),
                'org':           row.get('Organización','').strip(),
                'colaboradoras': colabs,
                'departamentos': deptos,
                'municipios':    municipios,
                'sectores':      sectores,
            })
    return respuestas


def red_to_js(respuestas):
    lines = []
    for r in respuestas:
        props = [
            f"nombre:{j(r['nombre'])}",
            f"org:{j(r['org'])}",
            f"colaboradoras:{j(r['colaboradoras'])}",
            f"departamentos:{j(r['departamentos'])}",
            f"municipios:{j(r['municipios'])}",
            f"sectores:{j(r['sectores'])}",
        ]
        lines.append("    {" + ",".join(props) + "}")
    return "{\n  respuestas: [\n" + ",\n".join(lines) + "\n  ]\n}"


# ─── reemplazo en index.html ─────────────────────────────────────────────────
# Patrón: "const NOMBRE = [" hasta el "
# ];" en su propia línea (no greedy).

def replace_const(content, name, js_value):
    # Handles both array  const X = [...\n];
    # and       object    const X = {...\n};
    pattern  = r'const ' + re.escape(name) + r' = [\[{][\s\S]*?\n[}\]];'
    new_text = f'const {name} = {js_value};'
    new_content, n = re.subn(pattern, lambda m: new_text, content, count=1)
    if n == 0:
        print(f"  ⚠️  'const {name}' no encontrado en index.html — omitido.")
    return new_content


def update_html(padb_rows, orgs, personas, respuestas):
    with open(HTML_FILE, 'r', encoding='utf-8') as f:
        content = f.read()

    original = content
    content  = replace_const(content, 'DEFAULT_DATA',  padb_to_js(padb_rows))
    content  = replace_const(content, 'ORGS_DATA',     orgs_to_js(orgs))
    content  = replace_const(content, 'PERSONAS_DATA', personas_to_js(personas))
    content  = replace_const(content, 'RED_DATA',      red_to_js(respuestas))

    if content == original:
        print("  ℹ️  index.html no cambió (datos idénticos).")
        return

    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.write(content)

    print("  ✅ index.html actualizado.")
    print(f"     {len(padb_rows)} acciones PADB  |  "
          f"{len(orgs)} organizaciones  |  "
          f"{len(personas)} personas  |  "
          f"{len(respuestas)} respuestas red")


# ─── main ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("🌿  TNC Bioeconomía — Actualizador de datos\n")

    ok = True

    print("📖  raw/padb_eje_cafetero.xlsx ...")
    try:
        padb_rows = read_padb()
        print(f"  ✓  {len(padb_rows)} acciones")
    except FileNotFoundError:
        print(f"  ❌  No encontrado: {PADB_FILE}")
        ok = False; padb_rows = []
    except Exception as e:
        print(f"  ❌  {e}"); ok = False; padb_rows = []

    print("📖  raw/actores.xlsx ...")
    try:
        orgs, personas = read_actores()
        print(f"  ✓  {len(orgs)} organizaciones, {len(personas)} personas")
    except FileNotFoundError:
        print(f"  ❌  No encontrado: {ACTORES_FILE}")
        ok = False; orgs = []; personas = []
    except Exception as e:
        print(f"  ❌  {e}"); ok = False; orgs = []; personas = []

    print("📖  raw/resultados_red.csv ...")
    try:
        respuestas = read_red_csv()
        print(f"  ✓  {len(respuestas)} respuestas")
    except FileNotFoundError:
        print(f"  ⚠️  No encontrado: {RED_CSV} — RED_DATA no se actualizará.")
        respuestas = None
    except Exception as e:
        print(f"  ❌  {e}"); ok = False; respuestas = None

    if not ok:
        print("\n❌  Abortado."); sys.exit(1)

    if respuestas is None:
        respuestas = []   # skip RED_DATA update gracefully

    print("\n✍️   Actualizando index.html ...")
    update_html(padb_rows, orgs, personas, respuestas)
    print("\n🎉  Listo. Recarga el navegador para ver los cambios.")
